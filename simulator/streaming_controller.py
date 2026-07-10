"""
streaming_controller.py
-----------------------
Manages the Server-Sent Events (SSE) stream for a single engine session.

Usage:
  controller = StreamingController()
  asyncio.create_task(engine.run(controller))   # engine pushes events
  return StreamingResponse(controller.stream(), media_type="text/event-stream")
"""

import asyncio
import json
import logging
import os
from collections import OrderedDict
from datetime import datetime
from typing import AsyncGenerator, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

_HEARTBEAT_INTERVAL = 5.0   # seconds between keep-alive pings
_QUEUE_TIMEOUT = 1.0        # seconds to wait before sending heartbeat

# Directory where XLSX reports are saved (relative to project root)
_REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")

# Events excluded from recording.
# `minute_pnl` stays on the live SSE stream, but skipping it here keeps
# long backtests from accumulating huge in-memory XLSX payloads.
_CSV_SKIP_EVENTS = {"monitor", "minute_pnl"}

# Position state fields — carried forward into every event row from the last known state
_POSITION_STATE_KEYS = {
    "lot", "lot_size", "quantity",
    "ce_strike", "ce_entry_price", "ce_ltp", "ce_pnl",
    "pe_strike", "pe_entry_price", "pe_ltp", "pe_pnl",
    "ce_hedge_strike", "ce_hedge_entry_price", "ce_hedge_ltp", "ce_hedge_pnl",
    "pe_hedge_strike", "pe_hedge_entry_price", "pe_hedge_ltp", "pe_hedge_pnl",
}

# PnL state fields — carried forward so every event row shows latest PnL
_PNL_STATE_KEYS = {
    "pnl", "pnl_with_charges", "total_charges",
    "realized_charges", "estimated_active_charges",
    "cumulative_pnl", "cumulative_pnl_pct",
    "expiry_gross_pnl", "expiry_net_pnl", "expiry_total_charges",
}

# All columns in display order
_CSV_COLUMNS = [
    # ── Identity ──────────────────────────────────────────────────────
    "wall_clock_ts", "event", "timestamp",
    # ── Basic setup ───────────────────────────────────────────────────
    "lot", "lot_size", "quantity",
    "expiry", "current_expiry",
    # ── CE Sell position ──────────────────────────────────────────────
    "ce_strike", "ce_entry_price", "ce_ltp", "ce_pnl",
    # ── PE Sell position ──────────────────────────────────────────────
    "pe_strike", "pe_entry_price", "pe_ltp", "pe_pnl",
    # ── CE Hedge position ─────────────────────────────────────────────
    "ce_hedge_strike", "ce_hedge_entry_price", "ce_hedge_ltp", "ce_hedge_pnl",
    # ── PE Hedge position ─────────────────────────────────────────────
    "pe_hedge_strike", "pe_hedge_entry_price", "pe_hedge_ltp", "pe_hedge_pnl",
    # ── Market data ───────────────────────────────────────────────────
    "spot", "atm", "fifth_otm_premium", "strike_number",
    "initial_premium_collected", "margin_used",
    # ── PnL ───────────────────────────────────────────────────────────
    "pnl", "pnl_with_charges", "total_charges",
    "realized_charges", "estimated_active_charges",
    "cycle_pnl", "cycle_charges",
    "cumulative_pnl", "cumulative_pnl_pct",
    "expiry_gross_pnl", "expiry_net_pnl", "expiry_total_charges",
    # ── Close / reason ────────────────────────────────────────────────
    "reason",
    # ── Adjustment ────────────────────────────────────────────────────
    "side", "message",
    "spot_price", "upper_adjustment_price", "lower_adjustment_price",
    "option_type", "sold_strike", "current_atm", "required_shift",
    # ── Re-entry / scheduling ─────────────────────────────────────────
    "reentry_delay", "scheduled_for", "target_expiry",
    "previous_expiry", "next_expiry",
    # ── Lifecycle ─────────────────────────────────────────────────────
    "status",
]

# ── Excel styling constants ────────────────────────────────────────────────────

# Header: dark blue background, white bold text
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# Date separator row: amber background, black bold text
_DATE_SEP_FILL = PatternFill("solid", fgColor="FFE699")
_DATE_SEP_FONT = Font(bold=True, color="1A1A1A")

# Summary total row: dark green
_TOTAL_FILL = PatternFill("solid", fgColor="1E5631")
_TOTAL_FONT = Font(bold=True, color="FFFFFF")

# Events that are ALWAYS logged regardless of frequency
_ALWAYS_LOG_EVENTS = {
    "started", "stopped", "error",
    "positions_opened", "positions_closed",
    "hedge_opened", "hedge_closed",
    "adjustment_triggered", "otm_adjustment_triggered",
    "reentry_scheduled", "event_reentry_scheduled", "next_expiry_scheduled",
    "stoploss_hit", "target_hit", "trailing_sl_hit",
    "minute_pnl",
    "expiry_exit",
}


class StreamingController:
    def __init__(self, position_start_time: str = "09:15") -> None:
        self._queue: asyncio.Queue[str] = asyncio.Queue()
        self._stopped: bool = False

        # Hourly monitor log filter
        self._log_minute: str = position_start_time.split(":")[1]
        self._last_monitor_log_hhmm: str = ""

        # XLSX recording — collects all non-monitor events; written on "stopped"
        # Each entry: (expiry_tag: str, row: dict)
        self._csv_rows: list[tuple[str, dict]] = []

        # Carries last known position state forward into every event row
        self._position_state: dict = {k: "" for k in _POSITION_STATE_KEYS}

        # Carries last known PnL state forward into every event row
        self._pnl_state: dict = {k: "" for k in _PNL_STATE_KEYS}

        # Tracks which expiry the current events belong to (for sheet grouping)
        self._current_expiry_tag: str = "Unknown"

        # Stores the engine's authoritative backtest_summary for the Profit Summary sheet
        self._backtest_summary: Optional[dict] = None

        session_tag = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(_REPORTS_DIR, exist_ok=True)
        self._csv_path: str = os.path.join(_REPORTS_DIR, f"backtest_{session_tag}.xlsx")

    # ------------------------------------------------------------------
    # Push side (called by StrategyEngine)
    # ------------------------------------------------------------------

    async def send(self, event_type: str, data: dict) -> None:
        """
        Serialize and queue one SSE event.

        Logging rules:
          - All non-monitor events → always logged at INFO level.
          - monitor events         → logged only once per hour,
                                     at the minute matching position_start_time.

        CSV recording: all events except monitor are stored; flushed to disk on "stopped".
        """
        wall_ts = datetime.now().isoformat(timespec="milliseconds")
        payload = json.dumps({"event": event_type, "ts": wall_ts, "data": data})
        await self._queue.put(payload)

        # Capture authoritative summary from engine
        if event_type == "backtest_summary":
            self._backtest_summary = data

        # CSV — record every event except monitor
        if event_type not in _CSV_SKIP_EVENTS:
            self._record_csv_row(event_type, wall_ts, data)

        # Write XLSX when backtest finishes
        if event_type == "stopped":
            self._write_csv()

        # Logging — monitor is hourly; all others always logged
        if event_type == "monitor":
            self._maybe_log_monitor(data)
        elif event_type in _ALWAYS_LOG_EVENTS:
            logger.info(f"[{event_type}] {self._summarise(event_type, data)}")
        else:
            logger.debug(f"[stream] queued event={event_type}")

    def _maybe_log_monitor(self, data: dict) -> None:
        """Log a monitor event only when minute == position_start_time's minute."""
        ts: str = data.get("timestamp", "")
        if not ts:
            return

        sep = "T" if "T" in ts else " "
        hhmm = ts.split(sep)[-1][:5]
        current_minute = hhmm.split(":")[1]

        if current_minute != self._log_minute:
            return
        if hhmm == self._last_monitor_log_hhmm:
            return

        self._last_monitor_log_hhmm = hhmm
        spot   = data.get("spot", "?")
        pnl    = data.get("current_pnl", data.get("pnl", "?"))
        expiry = data.get("current_expiry", "?")
        logger.info(
            f"[monitor] {ts} | spot={spot} | pnl={pnl} | expiry={expiry}"
        )

    @staticmethod
    def _summarise(event_type: str, data: dict) -> str:
        """Build a concise one-line log string for important events."""
        ts = data.get("timestamp", "")
        if event_type == "positions_opened":
            return (
                f"{ts} | CE={data.get('ce_strike')}@{data.get('ce_entry_price')} "
                f"PE={data.get('pe_strike')}@{data.get('pe_entry_price')} "
                f"expiry={data.get('expiry')}"
            )
        if event_type == "positions_closed":
            return (
                f"{ts} | reason={data.get('reason')} "
                f"cycle_pnl={data.get('cycle_pnl', data.get('pnl', '?'))} "
                f"cumulative_pnl={data.get('cumulative_pnl', '?')}"
            )
        if event_type == "adjustment_triggered":
            return (
                f"{ts} | side={data.get('side')} spot={data.get('spot_price')} "
                f"upper={data.get('upper_adjustment_price')} "
                f"lower={data.get('lower_adjustment_price')}"
            )
        if event_type in ("stoploss_hit", "target_hit", "trailing_sl_hit"):
            return (
                f"{ts} | cycle_pnl={data.get('cycle_pnl', '?')} "
                f"cumulative_pnl={data.get('cumulative_pnl', '?')}"
            )
        if event_type == "minute_pnl":
            return (
                f"{ts} | expiry={data.get('current_expiry', '?')} "
                f"expiry_net={data.get('expiry_net_pnl', data.get('pnl_with_charges', '?'))} "
                f"expiry_charges={data.get('expiry_total_charges', data.get('total_charges', '?'))}"
            )
        if event_type == "next_expiry_scheduled":
            return f"{ts} | {data.get('previous_expiry')} → {data.get('next_expiry')}"
        if event_type == "event_reentry_scheduled":
            return (
                f"{ts} | reason={data.get('reason')} mode={data.get('event_hit_position_status')} "
                f"scheduled_for={data.get('scheduled_for')} expiry={data.get('target_expiry')}"
            )
        if event_type == "hedge_opened":
            return (
                f"{ts} | CE={data.get('ce_hedge_strike')}@{data.get('ce_hedge_price')} "
                f"PE={data.get('pe_hedge_strike')}@{data.get('pe_hedge_price')}"
            )
        return f"{ts} | {data}"

    # ------------------------------------------------------------------
    # CSV / XLSX helpers
    # ------------------------------------------------------------------

    def _record_csv_row(self, event_type: str, wall_ts: str, data: dict) -> None:
        """
        Flatten one event's data dict into a row.

        State merge priority (lowest → highest):
          1. _pnl_state      — latest known PnL (so every event shows PnL)
          2. _position_state — latest known strikes / prices
          3. this event's own fields
        """
        # ── Update expiry tracker (for sheet grouping) ─────────────────
        for key in ("current_expiry", "expiry"):
            val = data.get(key)
            if val:
                self._current_expiry_tag = str(val)
                break

        # ── Update position state ──────────────────────────────────────
        for key in _POSITION_STATE_KEYS:
            if key in data and data[key] != "":
                self._position_state[key] = data[key]

        # ── Update PnL state ──────────────────────────────────────────
        for key in _PNL_STATE_KEYS:
            if key in data and data[key] != "":
                self._pnl_state[key] = data[key]

        # ── Build row ─────────────────────────────────────────────────
        row: dict = {col: "" for col in _CSV_COLUMNS}
        row["wall_clock_ts"] = wall_ts
        row["event"] = event_type

        # 1. Merge latest PnL state (lowest priority)
        for key, value in self._pnl_state.items():
            if key in row:
                row[key] = value

        # 2. Merge latest position state (medium priority)
        for key, value in self._position_state.items():
            if key in row:
                row[key] = value

        # 3. Overlay this event's own fields (highest priority)
        # Skip empty strings for carry-forward keys so closed positions persist until reset
        _carry_forward_keys = _POSITION_STATE_KEYS | _PNL_STATE_KEYS
        for key, value in data.items():
            if key in row:
                if value == "" and key in _carry_forward_keys:
                    continue
                row[key] = json.dumps(value) if isinstance(value, (dict, list)) else value

        self._csv_rows.append((self._current_expiry_tag, row))

    # ------------------------------------------------------------------
    # XLSX writer — expiry-wise sheets + date separator rows + summary
    # ------------------------------------------------------------------

    def _write_csv(self) -> None:
        """
        Write all collected rows to an XLSX file.

        Sheet layout:
          • One sheet per expiry (named by expiry date)
          • Each sheet has a frozen header row
          • A coloured date-separator row is inserted whenever the trading date changes
          • Final sheet: "Profit Summary" with one row per expiry + overall total
        """
        try:
            wb = Workbook()
            wb.remove(wb.active)   # remove default blank sheet
            sample_limit = 250

            # ── Group rows by expiry ───────────────────────────────────
            expiry_rows: OrderedDict[str, list[dict]] = OrderedDict()
            for expiry_tag, row in self._csv_rows:
                expiry_rows.setdefault(expiry_tag, []).append(row)

            for expiry_tag, rows in expiry_rows.items():
                sheet_name = str(expiry_tag)[:31]
                ws = wb.create_sheet(title=sheet_name)

                # Header row
                ws.append(_CSV_COLUMNS)
                for cell in ws[1]:
                    cell.font      = _HEADER_FONT
                    cell.fill      = _HEADER_FILL
                    cell.alignment = _HEADER_ALIGN
                ws.freeze_panes = "A2"

                # Data rows with date separators
                current_date: Optional[str] = None
                for row in rows:
                    ts = row.get("timestamp", "") or row.get("wall_clock_ts", "")
                    row_date = ts[:10] if ts else ""

                    if row_date and row_date != current_date:
                        current_date = row_date
                        # Coloured date-separator row
                        sep_row = [""] * len(_CSV_COLUMNS)
                        sep_row[0] = f"  {row_date}  "
                        ws.append(sep_row)
                        sep_excel_row = ws[ws.max_row]
                        for cell in sep_excel_row:
                            cell.fill      = _DATE_SEP_FILL
                            cell.font      = _DATE_SEP_FONT
                            cell.alignment = _HEADER_ALIGN

                    ws.append([row.get(col, "") for col in _CSV_COLUMNS])

                # Auto-fit column widths (capped at 40)
                for col_idx, col_name in enumerate(_CSV_COLUMNS, start=1):
                    sampled_values = [col_name]
                    for row in rows[:sample_limit]:
                        sampled_values.append(str(row.get(col_name, "") or ""))
                    max_len = max((len(value) for value in sampled_values), default=10)
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                        max_len + 2, 40
                    )

            # ── Profit Summary sheet (uses engine's authoritative backtest_summary) ──
            ws_sum = wb.create_sheet(title="Profit Summary")

            if self._backtest_summary:
                s = self._backtest_summary

                # ── Section 1: Per-expiry breakdown ───────────────────
                expiry_cols = [
                    "Expiry", "Trade Cycles", "Winning Trades", "Losing Trades",
                    "Win Rate %", "Gross PnL", "Total Charges", "Net PnL",
                    "First Entry", "Last Exit",
                ]
                ws_sum.append(expiry_cols)
                for cell in ws_sum[1]:
                    cell.font      = _HEADER_FONT
                    cell.fill      = _HEADER_FILL
                    cell.alignment = _HEADER_ALIGN
                ws_sum.freeze_panes = "A2"

                cumulative = 0.0
                for eb in s.get("expiry_breakdown", []):
                    net = eb.get("net_pnl", 0) or 0
                    cumulative = round(cumulative + net, 2)
                    cycles = eb.get("trade_cycles", 0) or 0
                    wins   = eb.get("winning_trades", 0) or 0
                    win_rate = round(wins / cycles * 100, 2) if cycles else 0.0
                    ws_sum.append([
                        eb.get("expiry", ""),
                        cycles,
                        wins,
                        eb.get("losing_trades", 0),
                        win_rate,
                        eb.get("gross_pnl", ""),
                        eb.get("total_charges", ""),
                        net,
                        eb.get("first_entry_time", ""),
                        eb.get("last_exit_time", ""),
                    ])

                # Total row for expiry section
                ws_sum.append([
                    "TOTAL",
                    s.get("total_trades", ""),
                    s.get("total_winning_trades", ""),
                    s.get("total_losing_trades", ""),
                    s.get("win_rate", ""),
                    round((s.get("total_profit", 0) or 0) + (s.get("total_loss", 0) or 0), 2),
                    s.get("total_charges", ""),
                    s.get("net_pnl", ""),
                    "", "",
                ])
                for cell in ws_sum[ws_sum.max_row]:
                    cell.font      = _TOTAL_FONT
                    cell.fill      = _TOTAL_FILL
                    cell.alignment = _HEADER_ALIGN

                # ── Section 2: Overall stats (2 blank rows gap) ───────
                ws_sum.append([])
                ws_sum.append([])

                stat_header = ["Metric", "Value"]
                ws_sum.append(stat_header)
                for cell in ws_sum[ws_sum.max_row]:
                    cell.font      = _HEADER_FONT
                    cell.fill      = _HEADER_FILL
                    cell.alignment = _HEADER_ALIGN

                stats = [
                    ("Total Trades",             s.get("total_trades", "")),
                    ("Winning Trades",            s.get("total_winning_trades", "")),
                    ("Losing Trades",             s.get("total_losing_trades", "")),
                    ("Win Rate %",                s.get("win_rate", "")),
                    ("Total Profit",              s.get("total_profit", "")),
                    ("Total Loss",                s.get("total_loss", "")),
                    ("Total Charges",             s.get("total_charges", "")),
                    ("Net PnL",                   s.get("net_pnl", "")),
                    ("Overall PnL %",             s.get("overall_pnl_pct", "")),
                    ("Max Profit (single cycle)", s.get("max_profit", "")),
                    ("Max Loss (single cycle)",   s.get("max_loss", "")),
                    ("Max Drawdown",              s.get("max_drawdown", "")),
                    ("Risk Reward Ratio",         s.get("risk_reward_ratio", "")),
                    ("Avg Profit / Trade",        s.get("average_profit_per_trade", "")),
                    ("Avg Loss / Trade",          s.get("average_loss_per_trade", "")),
                    ("Total Expiries Traded",     s.get("total_expiries_traded", "")),
                ]
                for metric, value in stats:
                    ws_sum.append([metric, value])

            else:
                # Fallback: no backtest_summary received — write a note
                ws_sum.append(["No backtest_summary data available"])

            # Auto-fit summary columns
            for col_cells in ws_sum.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws_sum.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

            wb.save(self._csv_path)
            logger.info(
                f"[xlsx] Report saved → {self._csv_path} "
                f"({len(self._csv_rows)} rows, {len(expiry_rows)} expiry sheet(s))"
            )

        except Exception as exc:
            logger.error(f"[xlsx] Failed to write report: {exc}")

    def stop(self) -> None:
        """Signal that no more events will be produced."""
        self._stopped = True

    # ------------------------------------------------------------------
    # Pull side (consumed by FastAPI StreamingResponse)
    # ------------------------------------------------------------------

    async def stream(self) -> AsyncGenerator[str, None]:
        """
        Yields SSE-formatted strings.
        Sends a heartbeat comment every _HEARTBEAT_INTERVAL seconds so
        the HTTP connection stays alive during quiet periods.
        """
        while True:
            try:
                message = await asyncio.wait_for(
                    self._queue.get(), timeout=_QUEUE_TIMEOUT
                )
                yield f"data: {message}\n\n"
            except asyncio.TimeoutError:
                if self._stopped and self._queue.empty():
                    break
                # Keep-alive comment (ignored by SSE parsers)
                yield f": heartbeat {datetime.now().isoformat(timespec='seconds')}\n\n"
            except Exception as exc:
                logger.error(f"[stream] unexpected error: {exc}")
                break
